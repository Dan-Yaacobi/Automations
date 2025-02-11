#other libraries to be used in this script
import os
from openpyxl import load_workbook,Workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill, Border, Side,Font
import calendar
import pandas as pd
#global variables
employees = {} #key = name, value = an array of availabilities(strings)

notes = {}

bright_colors = {
    "neon_green": [144, 238, 144],
    "neon_blue": [0, 255, 255],
    "neon_pink": [255, 0, 127],
    "turquoise": [0, 255, 127],
    "magenta": [255, 0, 255],
    "lime_green": [205, 205, 0],
    "lemon_yellow": [255, 247, 0],
    "orange": [255, 165, 0],
    "hot_pink": [255, 105, 180],
    "purple": [128, 0, 128],
    "cyan": [0, 255, 255],
    "chartreuse": [127, 255, 0],
    "coral": [255, 127, 80],
    "gold": [255, 215, 0],
    "silver": [192, 192, 192],
    }
color_names = [
    "neon_green",
    "neon_blue",
    "neon_pink",
    "turquoise",
    "magenta",
    "lime_green",
    "lemon_yellow",
    "orange",
    "hot_pink",
    "purple",
    "cyan",
    "chartreuse",
    "coral",
    "gold",
    "silver",
    "white"
]

def open_folder_dialog(string):

    def select_excel():
        nonlocal excel_path
        excel_path = filedialog.askopenfilename(
            initialdir= os.getcwd(),  # Initial directory (change as needed)
            title='Select Excel File',
            filetypes=[('Excel Files', '*.xlsx;*.xls;*csv'), ('All Files', '*.*')]
        )
        if excel_path:
            print(f"Selected {string}: {excel_path}")
            root.quit()  # Exit the loop when a folder is selected

        
        

    excel_path = None
    root = tk.Tk()
    root.title(f"{string} Picker")
    label = tk.Label(root, text= f"Please select a {string}:",font =("Arial",13))
    label.pack(pady=20)
    root.geometry("300x200+800+400")

    button = tk.Button(root, text=f"Select {string}", command=select_excel, width = 15, height = 5 )
    button.pack(pady=5,padx=5)
    
    while excel_path is None:
        root.mainloop()
        if excel_path is None:
            label.config(text=f"Please select a {string} (required):")
    root.destroy()
    return excel_path



def build_employs_dict(excel_path):
    data = pd.read_csv(excel_path)
    current_directory = os.getcwd()
    excel_file_path = os.path.join(current_directory, "new_file.xlsx")
    data.to_excel(excel_file_path, index=False)
    workbook = load_workbook(excel_file_path)
    # Access the active sheet 
    sheet = workbook.active   
    start_cell = sheet['B2'] # this assumes the first row looks like name,day_1,day_2... and the first column is submission date
    names_index = 3
    while(start_cell.value != None):
        employees[start_cell.value] = []
        name = start_cell.value
        for i in range(1,32):
            start_cell = start_cell.offset(column=1)
            employees[name].append(start_cell.value)
        notes[name] = start_cell.offset(column = 1).value
        start_cell = sheet['B'+str(names_index)]
        names_index += 1
        


def create_excel():
    font = Font(size=20)
    workbook = Workbook()
    sheet = workbook.active
    pinkish = 'FFF7CAAC'
    greyish = 'FFDADADA'
    row_fill = PatternFill(start_color = pinkish, end_color = pinkish, fill_type = 'solid')
    column_fill = PatternFill(start_color = greyish, end_color = greyish, fill_type = 'solid')

    
    #Coloring the first column in gray and adding morning\evening at start of each 10 grey rows
    #also leaving a row for the row colors indicating the dates
    current_cell = sheet['A2']
    first_row = current_cell
    last_row = current_cell
    
    border_style = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thin'),bottom=Side(style='thin'))
    border_style2 = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thin'))
    for i in range(5):
        for k in range(2):
            if k == 0:
                current_cell.value = "בוקר"
                current_cell.font = font
                first_row = current_cell.row
            if k == 1:
                first_row = current_cell.row + 1
                current_cell.value = "ערב"
                current_cell.font = font
                current_cell.border = border_style2
            for j in range(10):
                current_cell.fill = column_fill
                current_cell = current_cell.offset(row=1)
                
            last_row = current_cell.row
            for rows in range(first_row, last_row+1):
                cell = sheet.cell(row=rows, column =1)
                cell.border = border_style
        current_cell = current_cell.offset(row=1)
                
    #Coloring the rows according to non weekend days            
    start_cell = sheet['A1']
    start_cell.fill = row_fill
    year = datetime.now().year
    month = datetime.now().month + 1
    days_in_month = calendar.monthrange(year, month)[1]
    week = 1
    for i in range(1,days_in_month+1):
        day = get_date(year, month, i)
        if day != 4 and day != 5: #friday or saturday
            today = str(i) + "." + str(month)
            start_cell.fill = row_fill
            start_cell = start_cell.offset(column=1)
            start_cell.value = today
            start_cell.font = font
            start_cell.fill = row_fill

        elif day == 4:
            if i > 1:
                start_cell = sheet['A'+str(week*21 + 1)]
                week += 1
            else:
                continue
        else:
            continue        
        
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 200
    for i in range(107):
        sheet.row_dimensions[i].height = 30
    print("Finished creating the calendar excel file")
    workbook.save("משמרות סופי.xlsx")
    
    
def get_date(year,month,day):
    date_object = datetime(year, month, day)
    return date_object.weekday()
    
def make_shifts_table():
    Avails = ['בוקר','ערב','שניהם','לא יכול/ה']
    font = Font(size=30)
    
    excel_path = open_folder_dialog('excel')
    build_employs_dict(excel_path)
    remove_nones()
    workbook = load_workbook(os.path.join(os.getcwd(),"משמרות סופי.xlsx"))
    sheet = workbook.active
    
    pinkish = 'FFF7CAAC'
    row_fill = PatternFill(start_color = pinkish, end_color = pinkish, fill_type = 'solid')
    notes_cell = sheet['H110']
    notes_cell.value = "הערות"
    notes_cell.font = font
    notes_cell.fill = row_fill
    notes_cell.offset(column = 1).fill = row_fill
    color = 0
    for key in employees:
        random_color = generate_random_color(color)
        color += 1
        fill = PatternFill(start_color=random_color, end_color=random_color, fill_type="solid")
        
        write_notes(sheet,fill,key,notes_cell)
        
        week_index = 0
        date_pointer = sheet['B1']
        for answer in employees[key]:
            if date_pointer.value == None:
                week_index += 1
                date_pointer = sheet['B'+str(1 + week_index*21)]
            shifts_pointer = date_pointer.offset(row=1) # always one underneath date_pointer
            if answer != Avails[3]: # if the answer is "לא יכול" then notihing happens
                if answer == Avails[0]: # בוקר
                    find_fill_empty_cell(sheet,shifts_pointer,key,fill)
                if answer == Avails[1]: #ערב
                    shifts_pointer = date_pointer.offset(row=11) # finds the nearest evening
                    find_fill_empty_cell(sheet, shifts_pointer, key, fill)
                if answer == Avails[2]: #שניהם
                    find_fill_empty_cell(sheet,shifts_pointer,key,fill)
                    shifts_pointer = date_pointer.offset(row= 11) # finds the nearest evening
                    find_fill_empty_cell(sheet, shifts_pointer, key, fill)
            date_pointer = date_pointer.offset(column=1)
    workbook.save("משמרות סופי.xlsx")
    os.remove(os.getcwd()+"\\new_file.xlsx")
    return    



def write_notes(sheet,fill,key,notes_cell):
    for i in range(111,125):
        sheet.row_dimensions[i].height = 100
    font = Font(size=20)
    while notes_cell.value != None:
        notes_cell = notes_cell.offset(row=1)
    notes_cell.value = key
    notes_cell.font = font
    notes_cell.fill = fill
    notes_cell.offset(column=1).value = notes[key]
    notes_cell.offset(column=1).font = font

def find_fill_empty_cell(sheet,shifts_pointer,key,fill):
    font = Font(size=20)
    while(shifts_pointer.value != None):
        shifts_pointer = shifts_pointer.offset(row=1)
    shifts_pointer.value = key
    shifts_pointer.font = font
    shifts_pointer.fill = fill

def remove_nones():
    to_remove = None
    for key in employees:
        employees[key] = [x for x in employees[key] if x != to_remove]


def generate_random_color(color):
    the_color = color_names[color]
    red = bright_colors[the_color][0]
    green = bright_colors[the_color][1]
    blue = bright_colors[the_color][2]
    return f"{hex(red)[2:]:0>2}{hex(green)[2:]:0>2}{hex(blue)[2:]:0>2}"  # Combine RGB values

create_excel()
make_shifts_table()
finish = input("Press any key to exit: ")

    
    
    
    

    














