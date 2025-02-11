import win32com.client
#other libraries to be used in this script
import os
import fitz # PyMuPDF
from openpyxl import load_workbook
import datetime
import shutil
from tkinter import simpledialog
import tkinter as tk
from tkinter import filedialog



#Global Variables:

finished_dictionary = {}
cant_find_course_dictionary = {}


# deletes all the emails we worked on 
def delete_done_msg_files(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".msg"):
            file_path = os.path.join(folder_path, file_name)
            os.remove(file_path)


#given a file_path and a string, will move the file to the folder with the string's name (or create the folder)
def move_to_folder(mail_path, string):
    new_folder_name = string
    new_folder_path = os.path.join(folder_path,new_folder_name)
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)
    shutil.copy(mail_path, new_folder_path)
    
def find_date(sheet):
    found_cell = None
    for column in sheet.iter_cols():
        cell_value = column[0].value
        if isinstance(cell_value, datetime.datetime) and cell_value == date:
            found_cell = column[0]
            break
    return found_cell

def update_excel_with_unfound_course_num(date,excel_path,course_num,page_num,course_name):
    workbook = load_workbook(excel_path)
    # Access the active sheet (you may need to modify this depending on your Excel file structure)
    sheet = workbook.active
    
    # Find the cell containing the search string
    found_cell = find_date(sheet)
    
    current_cell = sheet.cell(row=found_cell.row, column=found_cell.column + 1)
    while current_cell.value is not None:
        current_cell = sheet.cell(row=current_cell.row + 1, column=current_cell.column)    
    current_cell.value = course_num
    current_cell = sheet.cell(row=current_cell.row, column=current_cell.column + 3)
    current_cell.value = page_num
    current_cell = sheet.cell(row=current_cell.row, column=current_cell.column -2)
    current_cell.value = course_name
    
    workbook.save(excel_path)


#given the excel path, the date, the course number and page number:
    #the function will open the excel file, find the date, and update the page number in the corrosopnding course number
def update_excel(date,excel_path,course_num,page_num,course_name):
    # Load the workbook
    print(f"this is course name {course_name}")
    if page_num == 0:
        print("There is nothing to update")
        return
    print(f"################# Updateing excel with {page_num} papers on course {course_num} for {date.month} / {date.year} #################")
    workbook = load_workbook(excel_path)
    # Access the active sheet (you may need to modify this depending on your Excel file structure)
    sheet = workbook.active
    
    # Find the cell containing the search string
    found_cell = find_date(sheet)
    # If the search string was found, perform the operations
    if found_cell:
        # Move 1 cell to the left
        left_cell = sheet.cell(row=found_cell.row, column=found_cell.column + 1)
        
        # Find course number going down cells
        course_num_cell = None
        for row_index in range(left_cell.row, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=left_cell.column)
            if cell.value == course_num:
                course_num_cell = cell
                break
        
        # If n1 is found, move 3 cells to the left and update the value with n2
        if course_num_cell:
            update_cell = sheet.cell(row=course_num_cell.row, column=course_num_cell.column + 3)
            update_cell_2 = sheet.cell(row=course_num_cell.row, column=course_num_cell.column + 1)
            current_value = update_cell.value
            update_cell_2.value = course_name
            if current_value is None:
                current_value = 0
            update_cell.value = current_value + page_num
            # Save the changes
            workbook.save(excel_path)
            if course_num in finished_dictionary:
                finished_dictionary[course_num] += page_num
            else:
                finished_dictionary[course_num] = page_num
            print("################# Updated Excel file successfully. #################")
        else:
            if course_num in cant_find_course_dictionary:
                cant_find_course_dictionary[course_num][0] += page_num
                cant_find_course_dictionary[course_num][1] = course_name
            else:
                cant_find_course_dictionary[course_num] = []
                cant_find_course_dictionary[course_num].append(page_num)
                cant_find_course_dictionary[course_num].append(course_name)
            print("Couldn't find course number in the excel file")         
    else:
        print(f"the date '{date}' was not found in the Excel file.")


#given the message and a string:
    #the function will find the number associated with the string inside the message. 
    #a number is considered associated with the string if inside the message we have "string: 1234" 
    #for example if string = "סעיף תקציבי" and inside the message we have "סעיף תקציבי: 50" then 50 will be returned
def get_number_from_body(string, message):
    body = message.Body
    lines = body.splitlines()
    search_string = string
    found = False
    for line in lines:
        if search_string in line:
            found = True
            colon_index = line.find(":")
            if colon_index != -1:
                number = ""
                i = 1
                while (line[colon_index + i].isdigit() == False):
                    i+=1
                while(line[colon_index + i].isdigit()):
                    number += line[colon_index + i]
                    i+=1
                    if colon_index + i == len(line):
                        break
    if found:
        return number
    else:
        print(f"could not find {string} inside the message")
        return -1
def get_string_from_body(substring, message):
    body = message.Body
    lines = body.splitlines()
    search_string = substring
    found = False
    found_string = ""
    
    for line in lines:
        if search_string in line:
            found = True
            colon_index = line.find(":")
            if colon_index != -1:
                found_string = line[colon_index + 1:].strip()
                break  # Found the string, no need to continue searching
    
    if found:
        return found_string
    else:
        print(f"Could not find '{substring}' inside the message")
        return None
    
def get_body_text(message):
    course_num = get_number_from_body("סעיף תקציבי", message)
    participants_num = get_number_from_body(("מספר עותקים"), message)
    course_name = get_string_from_body("שם התכנית",message)
    if ((course_num == -1) or (participants_num == -1)): # atleast one of them was not found = invalid mail format
        return -1,-1,course_name
    else:
        return int(course_num), int(participants_num), course_name
        
        
        
def access_attached_files(message, folder_path):
    #message = inbox.Items.GetFirst()
    none_pdf = False
    course_name = None
    attachments = message.Attachments
    pages_sum = 0 # sums the pages of all the attached files
    participants_num = 0 # the amount of participants
    course_num = 0 # the course number, סעיף תקציבי
    if attachments.Count > 0:
        for attachment in range(attachments.Count):
            file_name = attachments.Item(attachment + 1).FileName
            file_extension = file_name.split('.')[-1] if '.' in file_name else ''
            
            if file_extension.lower() == 'pdf':
                print(f"{file_name} is a PDF file.")
                # Perform actions specific to PDF files
                course_num, participants_num, course_name = get_body_text(message) # finds in the body of the message the amount of pariticpants and the course number
                pages_sum += get_page_num(attachments.Item(attachment + 1),folder_path, "pdf")
            elif file_extension.lower() in ["doc","docx"]:
                print(f"{file_name} is a Word file.")
                none_pdf = True
            
    if ((course_num == -1) or (participants_num == -1)):
        return -1,-1,False,course_name
    else:
        return course_num, pages_sum * participants_num, none_pdf ,course_name# amount of pages multiplied by the amount of participants


def get_page_num(attachment,file_path, file_type):
    attachment_path = os.path.join(file_path, f"attachment.{file_type}")
    attachment.SaveAsFile(attachment_path)
    if file_type == "pdf":        
        # Use PyMuPDF to get the number of pages in the PDF
        pdf_document = fitz.open(attachment_path)
        num_pages = pdf_document.page_count
    
        # Close the PDF document
        pdf_document.close()
    
        
        # Delete the temporary file
    os.remove(attachment_path)
    
    return int(num_pages)


def get_mails_from_folder(folder_path,outlook,date,excel_path):
# Open each .msg file in the folder
    total_pages = 0
    course_num = 0
    found_none_pdf = False
    file_count = 0
    for file_name in os.listdir(folder_path):
        bad_format_flag = False
        if file_name.endswith(".msg"):
            file_count += 1
            print(f"################# Started working on '{file_name}' #################")
            # Construct the full file path
            file_path = os.path.join(folder_path, file_name)
            
            # Open the .msg file
            print(f"this is the fp {file_path}")
            msg = outlook.OpenSharedItem(file_path)
            course_num, total_pages, found_none_pdf, course_name = access_attached_files(msg, folder_path)
            if ((course_num == -1) or (total_pages == -1)):
                print("Email is in a bad format")
                move_to_folder(file_path, 'Emails_not_containing_course_information')
                bad_format_flag = True
            else:
                # Close the message to release resources
                update_excel(date,excel_path,course_num,total_pages,course_name)
            msg.Close(1)
           # time.sleep(1)
        else:
            continue
        if found_none_pdf:
            move_to_folder(file_path,'Emails_containing_non_pdf_files')
            print("there emails in the new folder containined files that needs to be handled manualy, NO NEED TO RECOUNT PDF FILES")
        elif bad_format_flag == False:
            move_to_folder(file_path,'Finished_emails')

    print(f"################# Finished working on '{file_name}' #################\n")
    for key in cant_find_course_dictionary:
        finished_dictionary[key] = cant_find_course_dictionary[key]
    for key in finished_dictionary:
        print(f"updated course {key} on date {date.month}/{date.year} with {finished_dictionary[key]} papers\n")
        
              
    print(f"total mails worked on: {file_count}")
    print(f"################# Finished working on {file_count} files at {folder_path} #################")

    

def open_folder_dialog(string):

    def select_excel():
        nonlocal excel_path
        excel_path = filedialog.askopenfilename(
            initialdir= os.getcwd(),  # Initial directory (change as needed)
            title='Select Excel File',
            filetypes=[('Excel Files', '*.xlsx;*.xls'), ('All Files', '*.*')]
        )
        if excel_path:
            print(f"Selected {string}: {excel_path}")
            root.quit()  # Exit the loop when a folder is selected

        
        

    excel_path = None
    root = tk.Tk()
    root.title(f"{string} Picker")
    label = tk.Label(root, text= f"Please select a {string}:",font =("Arial",13))
    label.pack(pady=20)
    root.geometry("300x200+800+400")

    button = tk.Button(root, text=f"Select {string}", command=select_excel, width = 15, height = 5 )
    button.pack(pady=5,padx=5)
    
    while excel_path is None:
        root.mainloop()
        if excel_path is None:
            label.config(text=f"Please select a {string} (required):")
    root.destroy()
    return excel_path


      # Return the selected folder path after the window is closed

def get_year_and_month():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    root.geometry("300x200")
    # Prompt for year input
    year = simpledialog.askinteger("Enter Year", "Please enter the year:")
    
    # Check if year is entered (and not canceled)
    if year is not None:
        # Prompt for month input
        month = simpledialog.askinteger("Enter Month", "Please enter the month (1-12):")     
        # Check if month is entered (and not canceled)
        if month is not None and 1 <= month <= 12:
            print(f"Year entered: {year}, Month entered: {month}")
            date = datetime.datetime(int(year),int(month), 1)
            root.destroy()
            return date
        else:
            print("Invalid month or operation canceled.")
            return -1
    else:
        print("Operation canceled.")



def create_backup(file_path):
    try:
        # Define the backup file name (you can modify this as needed)
        file_name, file_extension = os.path.splitext(file_path)

        # Define the backup file name
        backup_file = f"{file_name}_backup{file_extension}"
        
        # Copy the file to create a backup
        shutil.copy2(file_path, backup_file)
    
        print(f"Backup created successfully: {backup_file}")
        return backup_file
    except Exception as e:
        print(f"Backup creation failed: {e}")
        
        
        
        


################# Main Function #################
    

while(True):
    date = get_year_and_month()
    if date == -1:
        continue
    excel_path = open_folder_dialog("excel file")
    backup_file = create_backup(excel_path)
    folder_path = os.getcwd()
    
    outlook = win32com.client.Dispatch('outlook.application')
    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)
    get_mails_from_folder(folder_path,namespace,date,excel_path)
    for key in cant_find_course_dictionary:
        update_excel_with_unfound_course_num(date,excel_path,key,cant_find_course_dictionary[key][0],cant_find_course_dictionary[key][1])
    delete_done_msg_files(folder_path)
    finished = input("Press any key to exit: ")
    break






