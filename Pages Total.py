from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog
import os



#gets the excel:
def open_folder_dialog(string):

    def select_excel():
        nonlocal excel_path
        excel_path = filedialog.askopenfilename(
            initialdir= os.getcwd(),  # Initial directory (change as needed)
            title='Select Excel File',
            filetypes=[('Excel Files', '*.xlsx;*.xls'), ('All Files', '*.*')]
        )
        if excel_path:
            print(f"Selected {string}: {excel_path}")
            root.quit()  # Exit the loop when a folder is selected

        
        

    excel_path = None
    root = tk.Tk()
    root.title(f"{string} Picker")
    label = tk.Label(root, text= f"Please select a {string}:",font =("Arial",13))
    label.pack(pady=20)
    root.geometry("300x200+800+400")

    button = tk.Button(root, text=f"Select {string}", command=select_excel, width = 15, height = 5 )
    button.pack(pady=5,padx=5)
    
    while excel_path is None:
        root.mainloop()
        if excel_path is None:
            label.config(text=f"Please select a {string} (required):")
    root.destroy()
    return excel_path

def find_course_num(excel_path):
    workbook = load_workbook(excel_path)
    # Access the active sheet (you may need to modify this depending on your Excel file structure)
    sheet = workbook.active
    i = 2
    for col in sheet.iter_cols():
        if (i-2)%5 != 0:
            i += 1
        else:
            cell = sheet.cell(row = 1, column = i)
            paper_cell = sheet.cell(row = cell.row, column = cell.column + 3)
            while(cell.value != None):
                course_name = sheet.cell(row = paper_cell.row, column = paper_cell.column -2).value
                if cell.value in paper_dict:
                    paper_dict[cell.value][0] += int(paper_cell.value)
                else:
                    paper_dict[cell.value] = []
                    paper_dict[cell.value].append(int(paper_cell.value))
                    paper_dict[cell.value].append(course_name)
                cell = sheet.cell(row=cell.row + 1, column = cell.column)
                paper_cell = sheet.cell(row = cell.row, column = cell.column + 3) 
            i += 1
                
def write_dict_to_excel(data_dict, file_name):
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Write keys and values into adjacent columns
    for idx, (key, value) in enumerate(data_dict.items(), start=1):
        sheet.cell(row=idx, column=1).value = key
        sheet.cell(row=idx, column=2).value = value[1]
        sheet.cell(row=idx, column=3).value = value[0]

    # Save the workbook
    workbook.save(file_name)

################# Main #################

paper_dict = {}
excel_path = open_folder_dialog("excel file")
find_course_num(excel_path)
write_dict_to_excel(paper_dict, "סיכום.xlsx")
finished = input("press any key exit:")
